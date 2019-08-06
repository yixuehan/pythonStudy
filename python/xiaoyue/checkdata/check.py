#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill
import time
# import numpy as np
import os
import json
# import shutil


class Record:
    def __init__(self, line, _time, phone, empll_id,
                 empll_button_id,
                 work_path, work_path_id, work_path_name,
                 work_button_id, event_name, event_button_id):
        self.line = int(line)
        self.datatime = time.strptime(_time, '%Y/%m/%d %H:%M:%S')
        self.time = time.mktime(self.datatime)
        self.phone = phone
        self.empll_id = empll_id
        self.empll_button_id = empll_button_id
        self.work_path = work_path
        self.work_path_id = work_path_id
        self.work_path_name = work_path_name
        self.work_button_id = work_button_id
        self.event_name = event_name
        self.event_button_id = event_button_id

    def __eq__(self, that):
        return self.work_path_id == that.work_path_id

    def __str__(self):
        return '%s %s %s %s %s %s %s %s %s %s %s %s' % (
                self.line,
                self.datatime,
                self.time,
                self.phone,
                self.empll_id,
                self.empll_button_id,
                self.work_path,
                self.work_path_id,
                self.work_path_name,
                self.work_button_id,
                self.event_name,
                self.event_button_id
                )


def time_eq(t1, t2):
    return (t1 - t2) <= 10 * 60


def is_night(t):
    return t.tm_hour >= 16 or t.tm_hour < 6


excel_sheet = None


indexlcyh = list(range(1, 15))
indexfcc = list(range(1, 10))

time20190715 = time.strptime('2019/07/15 18:00:00', '%Y/%m/%d %H:%M:%S')
time20190715 = time.mktime(time20190715)
indexfcc20190715 = list(range(1, 11))

indexfccliving = list(range(1, 6))


index_begin_lcyh = 2
index_end_lcyh = 1
time_duration_lcyh = 45 * 60
time_next_duration_lcyh = 60 * 60

index_begin_fcc = 1
index_end_fcc = 2
time_duration_fcc = 35 * 60
time_next_duration_fcc = 50 * 60

index_begin_fccliving = 1
index_end_fccliving = 5
time_duration_fccliving_day = 15 * 60
time_next_duration_fccliving_day = 60 * 60

time_duration_fccliving_night = 15 * 60
time_next_duration_fccliving_night = 30 * 60


max_time_duration = 1.5 * 60 * 60


repeat_color = PatternFill(fill_type='solid', fgColor='A2CD5A')
info_lose_color = PatternFill(fill_type='solid', fgColor='8A2BE2')
lose_color = PatternFill(fill_type='solid', fgColor='CD0000')
time_color = PatternFill(fill_type='solid', fgColor='CDCD00')
pos_t = 'K%d'

f = open("result.txt", "w")


def append_value(cell, value):
    if cell.value is None:
        cell.value = value
    else:
        cell.value = cell.value + value


def check_time(last_group, group,
               time_duration_day, time_next_duration_day,
               time_duration_night, time_next_duration_night):
    # check time

    if not len(last_group):
        return
    assert len(group)

    group_is_night = is_night(group[0].datatime)
    last_group_is_night = is_night(last_group[0].datatime)

    if not group_is_night:
        time_duration = time_duration_day
    else:
        time_duration = time_duration_night

    if not last_group_is_night:
        time_next_duration = time_next_duration_day
    else:
        time_next_duration = time_next_duration_night

    if group[-1].time - group[0].time < time_duration - 60:
        s = '行[%d ~ %d] 时间间隔不足[%ds]实际为[%fs]少了[%dm%fs]\n' % \
            (group[0].line, group[-1].line,
             time_duration, group[-1].time - group[0].time,
             int((time_duration - group[-1].time + group[0].time) / 60),
             (time_duration - group[-1].time + group[0].time) % 60
             )
        f.write(s)
        global pos_t, time_color
        pos = pos_t % group[-1].line
        append_value(excel_sheet[pos], s)
        excel_sheet[pos].fill = time_color

    if group[0].time - last_group[0].time < time_next_duration - 60:
        s = "行[%d ~ %d] 和上一组时间间隔不足[%ds]实际为[%fs]少了[%dm%fs]\n" % \
                (group[0].line, group[-1].line,
                 time_next_duration, group[0].time - last_group[0].time,
                 int((time_next_duration - group[0].time + last_group[0].time) / 60),
                 (time_next_duration - group[0].time + last_group[0].time) % 60)
        f.write(s)
        pos = pos_t % group[-1].line
        append_value(excel_sheet[pos], s)
        excel_sheet[pos].fill = time_color

    if group[0].time - last_group[0].time > max_time_duration:
        s = "行[%d ~ %d] 和上一组时间间隔超过[%ds]实际为[%fs]多了[%dm%fs]\n" % \
                (group[0].line, group[-1].line,
                 max_time_duration, group[0].time - last_group[0].time,
                 int((group[0].time - last_group[0].time - max_time_duration) / 60),
                 (group[0].time - last_group[0].time - max_time_duration) % 60
                 )
        f.write(s)
        pos = pos_t % group[-1].line
        append_value(excel_sheet[pos], s)
        excel_sheet[pos].fill = time_color

    last_group = group


def check_record(records, index_range, index_begin, index_end,
                 time_duration_day, time_next_duration_day,
                 time_duration_night, time_next_duration_night):
    if not len(records):
        return
    max_repeats = len(index_range)/2
    group = []
    last_group = []
    repeat_record = []
    repeats = 0
    last_record = records[0]
    group = [last_record]
    max_skip_line = 4
    for record in records[1:]:
        # print(str(record))
        if record not in group and not record.line - last_record.line > max_skip_line:
            group.append(record)
        else:
            repeats = repeats + 1
            if record in repeat_record:
                print("什么鬼? line[%d ~ %d]\n" %
                      (group[0].line, record.line))
                # assert False
            repeat_record.append(record)
            if repeats > max_repeats or record.work_path_id == index_begin or \
               record.line - last_record.line > max_skip_line:
                # print(record.work_path, record.work_path_id, record.line, repeats, max_repeats)
                lefts = [i for i in index_range
                         if i not in [record.work_path_id for record in group]]
                check_time(last_group, group, time_duration_day,
                           time_next_duration_day,
                           time_duration_night, time_next_duration_night)
                if record.work_path_id == index_begin:
                    if len(lefts) != 0:
                        s = "行[%d ~ %d] 缺少打点[%s]\n" % \
                                (group[0].line, last_record.line, lefts)
                        f.write(s)
                        global pos_t, lose_color
                        pos = pos_t % last_record.line
                        append_value(excel_sheet[pos], s)
                        excel_sheet[pos].fill = lose_color
                    if len(repeat_record[:-1]) != 0:
                        repeat_ids = [record.work_path_id for record in repeat_record[:-1]]
                        s = "行[%d ~ %d] 有重复[%s]\n" % \
                            (group[0].line,
                             last_record.line,
                             repeat_ids)
                        f.write(s)
                        global repeat_color
                        pos = pos_t % last_record.line
                        append_value(excel_sheet[pos], s)
                        excel_sheet[pos].fill = repeat_color
                    repeat_record = [repeat_record[-1]]
                else:
                    if len(lefts) != 0:
                        s = "重复过多，行[%d ~ %d] 缺少打点[%s]\n" % \
                                (group[0].line, group[-1].line, lefts)
                        f.write(s)
                        pos = pos_t % group[-1].line
                        append_value(excel_sheet[pos], s)
                        excel_sheet[pos].fill = repeat_color
                last_group = group
                group = repeat_record
                repeat_record = []
                repeats = 0
        last_record = record

    if len(group):
        if len(repeat_record) != 0:
            repeat_ids = [record.work_path_id for record in repeat_record]
            s = "行[%d ~ %d] 有重复[%s]\n" % \
                (group[0].line, last_record.line, repeat_ids)
            f.write(s)
            pos = pos_t % last_record.line
            append_value(excel_sheet[pos], s)
            excel_sheet[pos].fill = repeat_color
        lefts = [i for i in index_range
                 if i not in [record.work_path_id for record in group]]
        if len(lefts) != 0:
            s = "行[%d ~ %d] 缺少打点[%s]\n" % \
                    (group[0].line, last_record.line, lefts)
            f.write(s)
            pos = pos_t % last_record.line
            append_value(excel_sheet[pos], s)
            excel_sheet[pos].fill = lose_color

        check_time(last_group, group, time_duration_day,
                   time_next_duration_day,
                   time_duration_night, time_next_duration_night)


def unique_record(records):
    if not len(records):
        return records
    rs = []
    rs = [records[0]]
    for record in records[1:]:
        if rs[-1].work_path_id == record.work_path_id:
            s = "第[%d]行和第[%d]行巡检点序号重复[巡查点序号%d]\n" % \
                    (rs[-1].line, record.line, record.work_path_id)
            f.write(s)
            pos = pos_t % record.line
            append_value(excel_sheet[pos], s)
            excel_sheet[pos].fill = repeat_color
            continue
        rs.append(record)
    return rs


def save_array(filename, array):
    with open(filename, 'w') as f:
        json.dump(array, f, ensure_ascii=False)


def load_array(filename):
    with open(filename, 'r') as f:
        return json.load(f, object_pairs_hook=dict)


def gen_xlsl(excel_file, excel_out):
    excel_arr = excel_file + "_array"

    mtime_ori = os.path.getmtime(excel_file)
    mtime_arr = os.path.getmtime(excel_arr)

    lines = []
    records = []
    wb = openpyxl.load_workbook(excel_file)
    global excel_sheet
    excel_sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    assert excel_sheet is not None
    if not os.path.exists(excel_arr) or mtime_ori > mtime_arr:
        sheet_name = wb.sheetnames[0]
        sheet = wb.get_sheet_by_name(sheet_name)
        for i in range(2, sheet.max_row):
            # for i in range(2, 11):
            print('读取第[%d]条记录' % i)
            line = sheet[i]
            line = [i] + [value.value for value in line]
            assert len(line) == 11
            lines.append(line)
        # 保存数组
        save_array(excel_arr, lines)
    else:
        lines = load_array(excel_arr)

    for i in range(len(lines)):
        record = Record(*lines[i])
        # print('处理第[%d]行记录' % record.line)
        if not record.work_path or not record.work_path_id:
            f.write('第[%d]条记录不全，跳过\n' % record.line)
            pos = pos_t % record.line
            excel_sheet[pos] = '记录不全'
            excel_sheet[pos].fill = info_lose_color
            continue
        record.work_path_id = int(record.work_path_id)
        records.append(record)

    recordlcyh = [record for record in records
                  if record.work_path == '龙城一号施工现场']
    recordlcyh = unique_record(recordlcyh)

    recordfcc = [record for record in records
                 if record.work_path == '翡翠城施工现场' and record.time < time20190715]
    recordfcc = unique_record(recordfcc)

    recordfcc20190715 = [record for record in records
                         if record.work_path == '翡翠城施工现场' and record.time > time20190715]
    recordfcc20190715 = unique_record(recordfcc20190715)

    recordfccliving = [record for record in records
                       if record.work_path == '翡翠城生活区']
    recordfccliving = unique_record(recordfccliving)

    print(len(recordlcyh), len(recordfcc), len(recordfcc20190715), len(recordfccliving))

    check_record(recordlcyh, indexlcyh, index_begin_lcyh, index_end_lcyh,
                 time_duration_lcyh, time_next_duration_lcyh,
                 time_duration_lcyh, time_next_duration_lcyh)

    check_record(recordfcc, indexfcc, index_begin_fcc, index_end_fcc,
                 time_duration_fcc, time_next_duration_fcc,
                 time_duration_fcc, time_next_duration_fcc)

    check_record(recordfcc20190715, indexfcc20190715, index_begin_fcc, index_end_fcc,
                 time_duration_fcc, time_next_duration_fcc,
                 time_duration_fcc, time_next_duration_fcc)

    check_record(recordfccliving, indexfccliving, index_begin_fccliving, index_end_fccliving,
                 time_duration_fccliving_day,
                 time_next_duration_fccliving_day,
                 time_duration_fccliving_night,
                 time_next_duration_fccliving_night)

    wb.save(excel_out)


if __name__ == '__main__':
    excel_file = '巡检记录_20190701-20190729.xlsx'
    excel_out = '巡检记录_20190701-20190729_out.xlsx'
    gen_xlsl(excel_file, excel_out)
